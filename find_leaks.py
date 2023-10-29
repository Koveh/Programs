import os
import re
from openpyxl import load_workbook, Workbook

PATTERNS = {
    'OpenAI': r'sk-\w{32,50}',
    'YFinance': r'yf-\w{32}',
    'Telegram': r'\d{10}:[A-Za-z0-9-_]{35}',
    'Google': r'AIza\w{35}',
    'Deepl': r'[a-f0-9]{8}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{12}:[A-Za-z0-9]+',
    'FRED': 'insert your fred api key here',
    'Slack': 'insert your slack api key here or use the default', # r'xox[baprs]-[0-9a-zA-Z]{10,48}',
    'password': 'insert your password here',
    '1_password': 'insert your next here'
}
    
def on_startup():
    print('''Welcome to the API Key Management Utility.

1. Scan the directory to identify potential API key leaks. 
   - An Excel file will be generated with all detected API keys.
  
2. Choose the directory where you'd like to store the .env file.
  
3. Open the generated Excel file:
   - Mark rows under the column 'Mark API rows to add to .env' to indicate which API keys you want to store in the .env file.
   - Mark rows under the column 'Mark rows to change' if you'd like to replace the API keys in the source code with either a comment or an environment variable.
   
4. Bonus Features:
   - Identify empty files and oversized files.
   - Opt to delete the marked ones.

Usage Tips:
- If you've previously created an Excel file, simply specify its directory and type 'excel' when prompted.
- To make changes to the .env file, type '.env' after choosing your desired folder.

Please follow the on-screen instructions to proceed.
    ''')

    file_path = os.getcwd() # get the current directory
    file_path = move(file_path)
    while True:
        choice = input("""\n\n Please select an option by typing the corresponding number: 
1. Create an Excel file ('createExcel') - This will scan your project for potential API key leaks, empty files, large files, and empty folders.
2. Create .env file ('env') - Store marked API keys from Excel into a .env file.
3. Remove API Keys ('removeAPI') - Replace marked API keys in selected files with comments or environment variables.
4. Remove Empty Files ('empty') - Delete files that you've marked as empty in the Excel file.
5. Remove Large Files ('largeFiles') - Delete files that you've marked as large in the Excel file.
6. Remove Empty Folders ('emptyFolders') - Delete empty folders in the project.
Type 'exit' to quit the program.
                       """)
        
        if choice in ['1', 'createExcel']:
            create_excel_file_with_files(file_path)
            find_leaks(initial_path, PATTERNS)
            find_empty_files(file_path, initial_path)
            find_large_files(file_path, initial_path)
            
        elif choice in ['2', 'env', 'createEnv']:
            api_keys = extract_api_keys_from_excel(initial_path, PATTERNS)
            if api_keys:
                write_to_env_file(api_keys, file_path)

        elif choice in ['3', 'removeAPI']:
            api_keys = extract_api_keys_from_excel(initial_path, PATTERNS)
            if api_keys:
                update_python_files(api_keys, initial_path)

        elif choice in ['4', 'empty']:
            delete_marked_files(initial_path, 'Empty Files', 3)

        elif choice in ['5', 'largeFiles']:
            delete_marked_files(initial_path, 'Books', 4)

        elif choice in ['6', 'emptyFolders']:
            print("under construction")

        elif choice == 'exit':
            print("Exiting the program.")
            break
        else:
            print("Invalid choice. Please try again.")
    
    print("Thank you for using the API Key Management Utility.")


def extract_api_keys_from_excel(initial_path, patterns=PATTERNS):
    """ Extracts marked API keys from an Excel file and returns them in a dictionary. """
    api_keys = {}
    
    try:
        workbook = load_workbook(os.path.join(initial_path, "leak_file.xlsx"))
    except FileNotFoundError:
        print("Excel file not found.")
        return api_keys
    except PermissionError:
        print("Permission denied while reading Excel file.")
        return api_keys
    
    if 'Rows With APIs' not in workbook.sheetnames:
        print("Rows With APIs sheet not found.")
        return api_keys
    
    worksheet = workbook['Rows With APIs']
    
    print("Please mark the rows in the Excel file. if you did it, continue.")
    input("Press Enter to continue...")
    
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=7):
        
        # Assume the 'mark' is in the 7th column (column G)
        mark_cell = row[6]
        api_cell = row[4]
        
        if mark_cell.value is None:
            continue
                    
        if mark_cell.value:  # Check if the cell is not empty
            api_values = api_cell.value.split(', ')

            # Iterate through the patterns and find matches
            for pattern_name, pattern in patterns.items():
                modified_pattern = f'(\'?"?{pattern}\'?"?)'
                for api_value in api_values:
                    if re.fullmatch(modified_pattern, api_value):
                        if pattern_name not in api_keys:
                            api_keys[pattern_name] = []
                        if api_value not in api_keys[pattern_name]:
                            api_keys[pattern_name].append(api_value)
    
    return api_keys
    

            
def write_to_env_file(api_keys, file_path):
    """ Writes API keys to a .env file. """
    env_path = os.path.join(file_path, ".env")
    existing_keys = {}
    
    # Load existing keys if .env file already exists
    if os.path.exists(env_path):
        with open(env_path, "r") as f:
            for line in f:
                if "=" in line:
                    k, v = line.strip().split("=", 1)
                    existing_keys[k] = v
    try:
        with open(env_path, "a") as f:  # Open in append mode
            for service, keys in api_keys.items():
                # Replace dots and make uppercase
                service = service.replace('.', '_').upper()
                
                for key in keys:
                    index = 1  # Start index numbering from 1
                    new_key_name = f"{service}{index}"
                    
                    # Check for existing keys and increment index if found
                    while new_key_name in existing_keys and existing_keys[new_key_name] != key:
                        index += 1
                        new_key_name = f"{service}{index}"
                    
                    # If the key-value pair already exists, skip writing it again
                    if new_key_name in existing_keys and existing_keys[new_key_name] == key:
                        print(f"Key {new_key_name} already exists with the same value. Skipping...")
                        continue
                    
                    # Add the new key to the existing_keys dictionary
                    existing_keys[new_key_name] = key
                  
                    f.write(f'{new_key_name}="{key}"\n')
                    print(f"Added {new_key_name} to .env file.")
    except:
        print("Error while writing to .env file.")


def update_python_files(api_keys, initial_path):
    """ Updates Python files to use environment variables for API keys. """
    
    user_choice = input("Insert auto-loading .env code or a 'Insert API Key' comment? Type 'load_env' or 'comment': ")

    workbook = load_workbook(os.path.join(initial_path, "leak_file.xlsx"))
    worksheet = workbook['Rows With APIs']
    
    # Iterate through the rows in the Excel file
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=8):
       mark_cell = row[7] 
       
       if mark_cell.value:  # Check if the cell is not empty
            filepath = row[1].value  # Assume the file path is in the 2nd column
            
            if filepath and filepath.endswith('.py'):
                try:
                    with open(filepath, 'r+') as f:
                        print(f'Opening {filepath}')
                        content = f.read()
                        # Prepare the lines to be inserted
                        lines_to_insert = []

                        if user_choice == 'load_env':
                            if 'load_env_from_root()' not in content:
                                # path to the .env file 
                                lines_to_insert.extend([
                                    'import os',
                                    'from dotenv import load_dotenv',
                                    '',
                                    'def load_env_from_root():',
                                    '    path = os.path.dirname(os.path.abspath(__file__))',
                                    '    while path != "/":',
                                    '        env_path = os.path.join(path, \'.env\')',
                                    '        if os.path.exists(env_path):',
                                    '            load_dotenv(dotenv_path=env_path)',
                                    '            return',
                                    '        path = os.path.dirname(path)',
                                    '',
                                    'load_env_from_root()'
                                ])
                        elif user_choice == 'comment' and '# The program was automatically analyzed by the API Key Management Utility by Koveh.com.' not in content:
                            lines_to_insert.append('# The program was automatically analyzed by the API Key Management Utility by Koveh.com.')

                        # Combine lines to insert into a single string
                        insert_str = '\n'.join(lines_to_insert) + '\n' if lines_to_insert else ''
                        # Add the insert string at the beginning of the file content
                        content = insert_str + content
                        
                        # Replace API keys
                        for service, keys in api_keys.items():
                            service = service.replace('.', '_').upper()  # Replace dots and make uppercase
                            for i, key in enumerate(keys):
                                variable_name = f'"{service}{i+1}"'
                                replace_with = f'os.environ.get({variable_name})' if user_choice == 'load_env' else '"INSERT_YOUR_API_KEY"'
                                # content = content.replace(f'"{key}"', replace_with)
                                # content = content.replace(f"'{key}'"", replace_with)

                                # # Use regular expression to replace the key
                                key_pattern = re.compile(rf'"[a-zA-Z0-9_-]*{re.escape(key)}[a-zA-Z0-9_-]*"')  # Continues to match characters after the known part of the key
                                content = key_pattern.sub(replace_with, content)
                                
                                key_pattern_2 = re.compile(rf"'[a-zA-Z0-9_-]*{re.escape(key)}[a-zA-Z0-9_-]*'")  # Continues to match characters after the known part of the key
                                content = key_pattern_2.sub(replace_with, content)
                                
                                
                                
                                # Using compiled regex patterns for better performance
                                # double_quoted_key_pattern = re.compile(rf'"[a-zA-Z0-9_-]*{re.escape(key)}[a-zA-Z0-9_-]*"')
                                # single_quoted_key_pattern = re.compile(rf"'[a-zA-Z0-9_-]*{re.escape(key)}[a-zA-Z0-9_-]*'")

                                # # Replace double-quoted keys
                                # content = double_quoted_key_pattern.sub(replace_with, content)

                                # # Replace single-quoted keys
                                # content = single_quoted_key_pattern.sub(replace_with, content)
                                
                        f.seek(0) # Move the cursor to the beginning of the file
                        f.write(content) # Write the content to the file
                        f.truncate() # Truncate the file to the current cursor position

                        print(f'Updated {filepath}')
                except PermissionError:
                    print(f'Permission denied while reading{filepath}')
                except UnicodeDecodeError:
                    print(f'UnicodeDecodeError while reading {filepath}')
                except FileNotFoundError:
                    print(f'FileNotFoundError while reading {filepath}')
                except:
                    print(f'Error while reading {filepath}')
                    
                
def move(file_path):
    show_files_folders(file_path)
    action = input("cd folder_name or cd ../ to move to parent directory. Press enter to choose the current directory: \n")
    if action == "cd ../":
        os.chdir("..")
        return move(os.getcwd())  # Return the new path to the original caller
    elif action.startswith("cd "):
        new_folder = action[3:]
        if new_folder in os.listdir(file_path):
            os.chdir(new_folder)
            return move(os.getcwd())  # Return the new path to the original caller
        else:
            for folder in os.listdir(file_path):
                if folder.upper() == new_folder.upper():
                    os.chdir(folder)
                    return os.getcwd()
            print("Folder does not exist.")
    elif action == "":
        return file_path
    else:
        return file_path

def find_leaks(file_path, patterns):
    print("find_leaks function is called.")
    count_apis = 0
    
    # # Load the workbook once
    workbook = load_workbook(initial_path + "/leak_file.xlsx")
    worksheet = workbook.active
    
    # # Create a new sheet for storing rows with APIs
    workbook.create_sheet('Rows With APIs')
    new_worksheet = workbook['Rows With APIs']
    
    # Create headers
    headers = [
        'Filename', 
        'File Path', 
        'Row Number', 
        'API Service', 
        'API Keys', 
        'Comments', 
        'Mark API rows to add to .env', 
        'Mark rows to change (add comment or variable from .env)'
    ]
    for idx, header in enumerate(headers, 1):
        new_worksheet.cell(row=1, column=idx, value=header)

    row_number = 2  # Start from the second row in the Excel sheet

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        file_name = row[0].value
        file_path = row[1].value
        program_names = []

        if not os.path.exists(file_path):
            print(f"{file_path} does not exist.")
            continue

        file_extension = file_name.split(".")[-1]

        try:
            with open(file_path, "r", encoding="utf-8") as f:
                data = f.read()
                matched_apis = []
                
                for program, pattern in patterns.items():
                    matches = re.findall(pattern, data)
                    if matches:   
                        matched_apis.extend(matches)
                        program_names.append(program)

                if matched_apis:
                    print("Before removing quotes:", matched_apis)
                
                    # check if ' or " in the match and remove ' and "
                    # for i, match in enumerate(matched_apis):
                    #     if match.startswith("'") or match.startswith('"'):
                    #         matched_apis[i] = match[1:]
                    #     if match.endswith("'") or match.endswith('"'):
                    #         matched_apis[i] = match[:-1]
                    
                    for i, match in enumerate(matched_apis):
                        new_match = match
                        if new_match.startswith("'") or new_match.startswith('"'):
                            new_match = new_match[1:]
                        if new_match.endswith("'") or new_match.endswith('"'):
                            new_match = new_match[:-1]
                        matched_apis[i] = new_match
                    print("Debugging matched_apis:", matched_apis)
                    
                    count_apis += 1
                    row[4].value = ', '.join(matched_apis)
                    row[5].value = ', '.join(program_names)

                    # Copy this row to the new sheet
                    for idx, cell in enumerate(row):
                        new_worksheet.cell(row=row_number, column=idx+1, value=cell.value)
                    
                    row_number += 1

        except UnicodeDecodeError:
            print(f"Skipping {file_path} due to encoding issues.")
    
    # Save the workbook once after all updates
    workbook.save(initial_path + "/leak_file.xlsx")

    print(f"The find_leaks function is finished. {count_apis} APIs were found.")


def find_empty_files(file_path, initial_path):
    print("Finding empty files.")
    
    # Ignore folders list
    ignore_folders = ['.git', 'venv']
    # Ignore file list
    ignore_files = ['__init__.py', '__init__.pyi']
    
    try:
        workbook = load_workbook(os.path.join(initial_path, "leak_file.xlsx"))
    except FileNotFoundError:
        print("Excel file not found.")
        return
    except PermissionError:
        print("Permission denied while reading Excel file.")
        return
    
    workbook.create_sheet('Empty Files')
    empty_worksheet = workbook['Empty Files']

    headers = [
        'Filename',
        'File Path',
        'Mark to delete'
    ]
    
    for idx, header in enumerate(headers, 1):
        empty_worksheet.cell(row=1, column=idx, value=header)

    row_idx = 2

    for root, _, filenames in os.walk(file_path):
        # Skip ignored folders
        if any(ignore_folder in root for ignore_folder in ignore_folders):
            continue

        for filename in filenames:
            if filename in ignore_files or filename.isupper():
                continue
            
            file_extension = filename.split(".")[-1]
            
            # Skip files with '.' in the first part of their names
            if '.' in filename.split('.')[0]:
                continue
            
             # Include files without an extension
            if '.' not in filename or file_extension in file_types:
                individual_file_path = os.path.join(root, filename)
                try:
                    if os.path.getsize(individual_file_path) == 0:
                        empty_worksheet.cell(row=row_idx, column=1, value=filename)
                        empty_worksheet.cell(row=row_idx, column=2, value=individual_file_path)
                        row_idx += 1
                except FileNotFoundError:
                    print(f"{individual_file_path} not found.")
                    continue
                except PermissionError:
                    print(f"Permission denied for {individual_file_path}.")
                    continue
    try:
        workbook.save(os.path.join(initial_path, "leak_file.xlsx"))
    except PermissionError:
        print("Permission denied while saving Excel file.")
        return
    
    print("Empty Files sheet has been created.")


def delete_marked_files(initial_path, sheet_name, mark_column):
    print(f"Deleting marked files from {sheet_name} sheet.")
    
    try:
        workbook = load_workbook(os.path.join(initial_path, "leak_file.xlsx"))
    except FileNotFoundError:
        print("Excel file not found.")
        return
    except PermissionError:
        print("Permission denied while reading Excel file.")
        return

    if sheet_name not in workbook.sheetnames:
        print(f"{sheet_name} sheet not found.")
        return

    worksheet = workbook[sheet_name]

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=mark_column):
        file_path_cell = row[1]
        # Assume the 'mark' is in the 3rd column (column C)
        delete_marker_cell = row[mark_column - 1]  # zero-based index

        # Check if the cell is not empty
        if delete_marker_cell.value:  
            file_path = file_path_cell.value
            try:
                os.remove(file_path)
                print(f"Deleted {file_path}.")
            except FileNotFoundError:
                print(f"{file_path} not found.")
            except PermissionError:
                print(f"Permission denied for {file_path}.")


def find_large_files(file_path, initial_path):
    print("Finding large files.")
    ignore_folders = ['.git', 'venv']
    # File types to analyze for being large files
    file_types = ["pdf", "epub", "fb2", "mobi", "azw3", "djvu", "doc", "docx", "rtf", "odt", "ppt", "pptx", "ods"]
    
    try:
        workbook = load_workbook(os.path.join(initial_path, "leak_file.xlsx"))
    except FileNotFoundError:
        print("Excel file not found.")
        return
    except PermissionError:
        print("Permission denied while reading Excel file.")
        return

    workbook.create_sheet('Books')
    books_worksheet = workbook['Books']

    book_headers = [
        'Filename', 
        'File Path', 
        'File Size (KB)', 
        'Mark to delete'
    ]

    for idx, header in enumerate(book_headers, 1):
        books_worksheet.cell(row=1, column=idx, value=header)
    
    row_idx = 2

    for root, _, filenames in os.walk(file_path):
        if any(ignore_folder in root for ignore_folder in ignore_folders):
            continue

        for filename in filenames:
            file_extension = filename.split(".")[-1]

            if file_extension.lower() in file_types:
                individual_file_path = os.path.join(root, filename)
                try:
                    file_size = os.path.getsize(individual_file_path) // 1024  # Convert bytes to KB
                    books_worksheet.cell(row=row_idx, column=1, value=filename)
                    books_worksheet.cell(row=row_idx, column=2, value=individual_file_path)
                    books_worksheet.cell(row=row_idx, column=3, value=file_size)
                    row_idx += 1
                except FileNotFoundError:
                    print(f"{individual_file_path} not found.")
                    continue
                except PermissionError:
                    print(f"Permission denied for {individual_file_path}.")
                    continue          
    try:
        workbook.save(os.path.join(initial_path, "leak_file.xlsx"))
    except PermissionError:
        print("Permission denied while saving Excel file.")
        return
    
    print("Books sheet has been created.")


def create_excel_file_with_files(file_path: str) -> None:
    """
    Create an Excel file with a list of files and their paths.

    Args:
        file_path (str): The path to the directory to search for files.

    Returns:
        None
    """
    print("creating leak excel file")
    
    excel_path = os.path.join(initial_path, "leak_file.xlsx")
    
    # check if the file exists
    if os.path.exists(excel_path):
        if input("leak_file.xlsx already exists. overwrite Y/N?: ") in ("Y", "y", "", "yes"):
            os.remove(excel_path)
        else:
            print("Aborting...")
            return

    
    total_files = 0
    files = []
    for root, folders, filenames in os.walk(file_path):
        for filename in filenames:
            if filename.split(".")[-1] not in file_types:
                continue
            total_files += 1
            individual_file_path = os.path.join(root, filename)  # Changed variable name here
            files.append({"file_name": filename, "file_path": individual_file_path})  # And here
    print(f"Total files: {total_files}")
    
    # No need to walk through the folders again as os.walk has already done that
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "All Files"
    headers = [
        'Filename', 
        'File Path', 
        'Row Number', 
        'API Service', 
        'API Keys',
        'List of APIs',
        'Program Names'
    ]
        
    for idx, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=idx, value=header)

    
    # write the data of the files, file paths
    row = 2
    for file in files:
        worksheet.cell(row=row, column=1, value=file["file_name"])
        worksheet.cell(row=row, column=2, value=file["file_path"])
        row += 1
    try:
        workbook.save(excel_path)
        print("leak_file.xlsx is created.")
    except PermissionError:
        print("Permission denied while saving Excel file.")
        return
    except:
        print("leak_file.xlsx is not created.")


def show_files_folders(file_path):
    """
    Display the files and folders in the given directory path.

    Args:
        file_path (str): The path of the directory to display.

    Returns:
        None
    """
    files = sorted([f for f in os.listdir(file_path) if os.path.isfile(os.path.join(file_path, f))])
    folders = sorted([f for f in os.listdir(file_path) if os.path.isdir(os.path.join(file_path, f))])
    
    if not os.listdir(file_path):
        print("The directory is empty")
        return
    print(f"Files and Folders in {file_path}:")
    for name in files:
        print(f"\033[92m{name}\033[0m")
    for name in folders:
        print(f"\033[94m{name}\033[0m")
    print("\n")

if __name__ == "__main__":
    initial_path = os.getcwd()
    #file_types = ["html", "csv", "txt", "py", "csv", "cs", "json", "env", "txt", "md", "js", "c", "html", "cpp"]
    file_types = ["py"]
       
    on_startup()


